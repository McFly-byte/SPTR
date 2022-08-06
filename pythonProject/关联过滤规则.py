import numpy as np
from numpy import *
import openpyxl

import openpyxl as op
import  pandas as pd

def huoqu(year1,year2):
    shuzi = []
    sheet = openpyxl.load_workbook('topic_cosine_lda_'+str(year1)+'_'+str(year2)+'.xlsx')
    sheet2 = []
    mySheet = sheet.get_sheet_by_name('Sheet1')
    col = mySheet["A"]
    for cell in col:
        sheet2.append(cell.value)
    sheet2.pop(0)
    sheet1 = []
    for i in sheet2:
        linshi1 = []
        a = i.replace('[','')
        b = a.replace(']','')
        linshi = b.split(',')
        for m in linshi:
            c = float(m)
            linshi1.append(c)
        sheet1.append(linshi1)
    juzhen = np.array(sheet1)
    hangzd = np.amax(juzhen,axis=1)
    liezd = np.amax(juzhen,axis=0)
    for i in hangzd:
        shuzi.append(i)
    for i in liezd:
        shuzi.append(i)
    return shuzi

#计算所有相似度平均值
def savage():
    shuzi = []
    for i in range(2008,2017,3):
        a = i+3
        linshi = huoqu(i,a)
        for m in linshi:
            shuzi.append(m)
    linshi1 = huoqu(2017,2019)
    for i in linshi1:
        shuzi.append(i)
    linshi2 = huoqu(2019,2021)
    for i in linshi2:
        shuzi.append(i)
    pingjunzhi = mean(shuzi)
    print(pingjunzhi)
    return pingjunzhi

#读取主题相似度矩阵
def duqu(year1,year2):
    sheet = openpyxl.load_workbook('topic_cosine_lda_'+str(year1)+'_'+str(year2)+'.xlsx')
    sheet2 = []
    mySheet = sheet.get_sheet_by_name('Sheet1')
    col = mySheet["A"]
    for cell in col:
        sheet2.append(cell.value)
    sheet2.pop(0)
    sheet1 = []
    for i in sheet2:
        linshi1 = []
        a = i.replace('[','')
        b = a.replace(']','')
        linshi = b.split(',')
        for m in linshi:
            c = float(m)
            linshi1.append(c)
        sheet1.append(linshi1)
    juzhen = np.array(sheet1)
    return juzhen

#关联过滤规则实现，不用看具体规则实现
def guize(year1,year2):
    juzhen = duqu(year1,year2)
    hangzd = np.amax(juzhen, axis=1)
    lie = juzhen.shape[1]
    yxztd = []
    pingjuzhi = savage()
    for i in range(len(hangzd)):
        if hangzd[i] >pingjuzhi:
            list_numpy_index_3 = np.where(juzhen == hangzd[i])
            list_numpy_index_3_matrix = np.dstack((list_numpy_index_3[0], list_numpy_index_3[1])).squeeze()
            qianzhuti = list_numpy_index_3_matrix[0]
            houzhuti = list_numpy_index_3_matrix[1]
            juzhen1 = juzhen[:,houzhuti].T
            paixu = np.argsort(-juzhen1)
            for a in range(5):
                if juzhen1[paixu[a]] < hangzd[i]:
                    if a == 4:
                        break
                    else:
                        continue
                else:
                    ls = str(year1)+'_topic'+str(qianzhuti+1)+'-'+str(year2)+'_topic'+str(houzhuti+1)+'-'+str(hangzd[i])
                    yxztd.append(ls)
                    if a == 0:
                        break
                    else:
                        for t in range(a):
                            list_index = np.where(juzhen == juzhen1[a])
                            list_index_matrix = np.dstack((list_index[0], list_index[1])).squeeze()
                            xinqianzhuti = list_index_matrix[0]
                            juzhen2 = juzhen[xinqianzhuti,:].T
                            paixu1 = np.argsort(-juzhen2)
                            if paixu1[0] == houzhuti:
                                ls = str(year1) + '_topic' + str(qianzhuti+1) + '-' + str(year2) + '_topic' + str(houzhuti+1) + '-' + str(hangzd[i])
                                yxztd.append(ls)
                            else:
                                if t == a-1:
                                    break
                                else:
                                    continue
    print(yxztd)
    return yxztd

def shuchu():
    bg = op.load_workbook(r"有效主题关联对.xlsx")  # 应先将excel文件放入到工作目录下
    sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下
    lie = 0
    for i in range(2008,2017,3):
        lie += 1
        a = i+3
        linshi = guize(i,a)
        for i in range(1, len(linshi) + 1):
            sheet.cell(i, lie, linshi[i - 1])  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
    linshi1 = guize(2017, 2019)
    lie += 1
    for i in range(1, len(linshi1) + 1):
        sheet.cell(i, lie, linshi1[i - 1])
    linshi2 = guize(2019, 2021)
    lie += 1
    for i in range(1, len(linshi2) + 1):
        sheet.cell(i, lie, linshi2[i - 1])
    bg.save("有效主题关联对.xlsx")  # 对文件进行保存
shuchu()