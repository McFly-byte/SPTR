# File: pythonProject/sankey
# user: mcfly
# IDE: PyCharm
# Create Time: 2022/8/18 14:51

import openpyxl
import pandas
from pyecharts.charts import  Sankey
from pyecharts import options as opts
import  pandas as pd

#读取
def dqzd():
    sheet = openpyxl.load_workbook('有效主题关联对.xlsx')
    mySheet = sheet.get_sheet_by_name('Sheet1')
    minrow = mySheet.min_row
    maxrow = mySheet.max_row
    mincol = mySheet.min_column
    maxcol = mySheet.max_column
    paixu = []
    for m in range(mincol, maxcol + 1):
        for n in range(minrow+1, maxrow + 1):
            cell = mySheet.cell(n, m).value
            paixu.append(str(cell))
    paixu1 = []
    for i in paixu:
        if i !='None':
            paixu1.append(i)
    return paixu1

#将数据转换为BDP软件画图的对应数据
def op_toExcel():  # openpyxl库储存数据到excel
    guanxi = dqzd()
    wb = openpyxl.Workbook()  # 创建工作簿对象
    ws = wb['Sheet']  # 创建子表
    ws.append(['source', 'target', 'value'])  # 添加表头
    for i in guanxi:
        a = i.split('-')
        d = a[0],a[1],a[2]
        ws.append(d)  # 每次写入一行
    wb.save('BDP数据.xlsx')

#定义桑基图节点
def nodes_():
    sheet = openpyxl.load_workbook('BDP数据.xlsx')
    sheet2 = []
    mySheet = sheet.get_sheet_by_name('Sheet')
    col = mySheet["A"]
    for cell in col:
        sheet2.append(cell.value)
    col1 = mySheet["B"]
    for cell in col1:
        sheet2.append(cell.value)
    sheet1 = list(set(sheet2))
    nodes = []
    for i in sheet1:
        dic = {}
        dic['name'] = i
        nodes.append(dic)
    return nodes



#定义桑基图关联关系
def links_():
    links = []
    sheet = pd.read_excel("BDP数据.xlsx", "Sheet")
    for row in sheet.index.values :
        dic = {}
        dic['source'] = str(sheet.iloc[row, 0])
        dic['target'] = str(sheet.iloc[row, 1])
        dic['value'] = float(sheet.iloc[row, 2])
        print(dic)
        links.append(dic)
    return links


#画出桑基图
def huatu(nodes,links):
    pic=(
        Sankey(init_opts=opts.InitOpts(width="1600px", height="700px")).add(
            '',#图例名称
            nodes,#传入节点数据
            links,#传入边和流量数据
            #设置线条透明度、弯曲度、颜色
            linestyle_opt=opts.LineStyleOpts(opacity=0.5,curve=0.5,color='source',),
            #标签显示位置
            label_opts=opts.LabelOpts(position='top'),
            #节点之间的距离
            node_gap=20,
            #设置节点可以拖动
            is_draggable= True,
            #设置节点
            itemstyle_opts=opts.ItemStyleOpts(border_width=0, border_color="#aaa"),
            #orient="vertical",#查看垂直图片的操作


        )
        .set_global_opts(title_opts=opts.TitleOpts(title='分时间窗下的主题相似度'))
    )
    pic.render('桑基图.html')

def sankey_diagram():
    op_toExcel()
    nodes = nodes_()
    print(nodes)
    links = links_()
    huatu(nodes,links)

if __name__ == '__main__':
    sankey_diagram()


